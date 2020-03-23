import re
from pathlib import Path
from zipfile import BadZipfile

import openpyxl
from dateutil import parser
from django.core.exceptions import ValidationError
from openpyxl import load_workbook
from phonenumber_field.validators import validate_international_phonenumber

from core.utils import get_choices_values
from core.validators import BaseValidator
from household.const import NATIONALITIES
from household.models import Individual, Household


class UploadXLSXValidator(BaseValidator):
    # TODO: Will be provided by utils.serialize_flex_attributes,
    #  temporarily hardcoded
    #  FLEX_ATTRS = serialize_flex_attributes()
    WB = None

    FLEX_ATTRS = {
        "individuals": {
            "id_type_i_f": {
                "type": "SELECT_ONE",
                "choices": (
                    "BIRTH_CERTIFICATE",
                    "DRIVERS_LICENSE",
                    "UNHCR_ID",
                    "NATIONAL_ID",
                    "NATIONAL_PASSPORT",
                    "OTHER",
                    "NOT_AVAILABLE",
                ),
            },
        },
        "households": {
            "assistance_type_h_f": {
                "type": "SELECT_MANY",
                "choices": (
                    "Option 1",
                    "Option 2",
                    "Option 3",
                    "Option 4",
                    "Option 5",
                    "Option 6",
                    "Option 7",
                ),
            },
            "water_source_h_f": {
                "type": "SELECT_ONE",
                "choices": (
                    "Buy bottled water",
                    "From piped water",
                    "From private vendor",
                    "To buy water from water tank",
                    "Collect water from rain water",
                    "Collect water from a well / source directly",
                ),
            },
        },
    }
    # TODO: Probably need to fetch fields directly from models
    CORE_FIELDS = {
        "individuals": {
            "household_id": {"type": "STRING"},
            "head_of_household": {
                "type": "SELECT_ONE",
                "choices": Individual.YES_NO_CHOICE,
            },
            "marital_status": {
                "type": "SELECT_ONE",
                "choices": Individual.MARTIAL_STATUS_CHOICE,
            },
            "status_as_head_of_household": {
                "type": "SELECT_ONE",
                "choices": ("ACTIVE", "N/A",),
            },
            "address": {"type": "STRING"},
            # TODO: We need to query locations and check,
            #  currently hardcoded
            #  those two fields are also missing in model fields
            "admin_level_1": {
                "type": "SELECT_ONE",
                "choices": ("AFGHANISTAN",),
            },
            "admin_level_2": {"type": "SELECT_ONE", "choices": ("KABUL",),},
            "phone_number_1": {"type": "PHONE_NUMBER"},
            "phone_number_2": {"type": "PHONE_NUMBER"},
            "given_name": {"type": "STRING"},
            "last_name": {"type": "STRING"},
            "middle_name": {"type": "STRING"},
            "full_name": {
                "type": "CALCULATED",
                "calculate_fields": ["first_name", "middle_name", "last_name",],
            },
            "sex": {"type": "SELECT_ONE", "choices": Individual.SEX_CHOICE,},
            "birth_date": {"type": "DATE"},
            "age": {"type": "CALCULATED", "calculate_fields": ["birth_date"],},
            "estimated_birth_date": {
                "type": "SELECT_ONE",
                "choices": Individual.YES_NO_CHOICE,
            },
            "work_status": {
                "type": "SELECT_ONE",
                "choices": Individual.YES_NO_CHOICE,
            },
            "disability": {
                "type": "SELECT_ONE",
                "choices": Individual.DISABILITY_CHOICE,
            },
            # TODO: this field is missing, temp. get it from file
            "severity_of_disability": {
                "type": "SELECT_ONE",
                "choices": ("A_LOT", "CANNOT_ALL", "SOME",),
            },
            "school_type": {
                "type": "SELECT_ONE",
                "choices": ("PUBLIC", "INFORMAL", "OTHER", "PRIVATE",),
            },
        },
        "households": {
            "household_id": {"type": "STRING"},
            "household_location": {"type": "GEOLOCATION"},
            "consent": {
                "type": "SELECT_ONE",
                "choices": Individual.YES_NO_CHOICE,
            },
            "residence_status": {
                "type": "SELECT_ONE",
                "choices": Household.RESIDENCE_STATUS_CHOICE,
            },
            "family_nationality": {
                "type": "SELECT_ONE",
                "choices": NATIONALITIES,
            },
            "household_size_h_c": {"type": "INTEGER"},
            "distance_from_school": {"type": "DECIMAL"},
        },
    }

    COMBINED_FIELDS_DICT = {
        **CORE_FIELDS["individuals"],
        **FLEX_ATTRS["individuals"],
        **CORE_FIELDS["households"],
        **FLEX_ATTRS["households"],
    }

    @classmethod
    def validate(cls, *args, **kwargs):
        validate_methods = [
            getattr(cls, m) for m in dir(cls) if m.startswith("validate_")
        ]

        errors_list = []
        for method in validate_methods:
            errors = method(cls, *args, **kwargs)
            errors_list.extend(errors)

        return errors_list

    @classmethod
    def string_validator(cls, value, *args, **kwargs):
        return isinstance(value, str)

    @classmethod
    def integer_validator(cls, value, *args, **kwargs):
        return isinstance(value, int)

    @classmethod
    def float_validator(cls, value, *args, **kwargs):
        return isinstance(value, float)

    @classmethod
    def geolocation_validator(cls, value, *args, **kwargs):
        pattern = re.compile(r"^(\-?\d+\.\d+?,\s*\-?\d+\.\d+?)$")
        return bool(re.match(pattern, value))

    @classmethod
    def date_validator(cls, value, *args, **kwargs):
        if cls.integer_validator(value):
            return False
        try:
            parser.parse(value)
        except ValueError:
            return False
        return True

    @classmethod
    def phone_validator(cls, value, *args, **kwargs):
        try:
            validate_international_phonenumber(value)
        except ValidationError:
            return False
        return True

    @classmethod
    def calculated_field_validator(cls, val, *args, **kwargs):
        return bool(val)

    @classmethod
    def choice_validator(cls, value, header, *args, **kwargs):
        choices = get_choices_values(
            cls.COMBINED_FIELDS_DICT[header]["choices"]
        )
        choice_type = cls.COMBINED_FIELDS_DICT[header]["type"]

        if choice_type == "SELECT_ONE":
            return value in choices
        elif choice_type == "SELECT_MANY":
            selected_choices = value.split(",")
            for choice in selected_choices:
                if choice.strip() not in choices:
                    return False
            return True

        return False

    @classmethod
    def not_empty_validator(cls, value, *args, **kwargs):
        return bool(value)

    @classmethod
    def rows_validator(cls, sheet):
        first_row = sheet[1]
        combined_fields = {
            **cls.CORE_FIELDS[sheet.title.lower()],
            **cls.FLEX_ATTRS[sheet.title.lower()],
        }

        switch_dict = {
            "STRING": cls.string_validator,
            "INTEGER": cls.integer_validator,
            "DECIMAL": cls.float_validator,
            "DATE": cls.date_validator,
            "DATETIME": cls.date_validator,
            "SELECT_ONE": cls.choice_validator,
            "SELECT_MANY": cls.choice_validator,
            "PHONE_NUMBER": cls.phone_validator,
            "CALCULATED": cls.not_empty_validator,
            "GEOLOCATION": cls.geolocation_validator,
            "GEOPOINT": cls.geolocation_validator,
            # TODO: add image validator, how image will be attached to file?
            # "IMAGE": cls.geolocation_validator,
        }

        invalid_rows = []
        for row in sheet.iter_rows(min_row=3):
            # openpyxl keeps iterating on empty rows so need to omit empty rows
            if not any([cell.value for cell in row]):
                continue

            for cell, header in zip(row, first_row):
                current_field = combined_fields[header.value]
                field_type = current_field["type"]
                fn = switch_dict.get(field_type)

                if fn(cell.value, header.value) is False:
                    message = (
                        "Unexpected value for type "
                        + f"{field_type.replace('_', ' ').lower()}"
                    )
                    invalid_rows.append(
                        {
                            "row_number": cell.row,
                            "header": header.value,
                            "message": message,
                        }
                    )

        return invalid_rows

    @classmethod
    def validate_file_extension(cls, *args, **kwargs):
        xlsx_file = kwargs.get("file")
        file_suffix = Path(xlsx_file.name).suffix
        if file_suffix != ".xlsx":
            return [
                {
                    "row_number": 1,
                    "header": f"{xlsx_file.name}",
                    "message": "Only .xlsx files are accepted for import.",
                }
            ]

        # Checking only extensions is not enough,
        # loading workbook to check if it is in fact true .xlsx file
        try:
            load_workbook(xlsx_file)
        except BadZipfile:
            return [
                {
                    "row_number": 1,
                    "header": "xlsx_file.nam",
                    "message": "Invalid .xlsx file",
                }
            ]

        return []

    @classmethod
    def validate_file_with_template(cls, *args, **kwargs):
        # TODO: temporarily check for a flex fields and core fields
        #  that are in template excel, have to check for all

        if cls.WB is None:
            xlsx_file = kwargs.get("file")
            wb = openpyxl.load_workbook(xlsx_file)
        else:
            wb = cls.WB

        for name, fields in cls.CORE_FIELDS.items():
            sheet = wb[name.capitalize()]
            first_row = sheet[1]

            expected_column_names = {
                *cls.CORE_FIELDS[name].keys(),
                *cls.FLEX_ATTRS[name].keys(),
            }
            column_names = {cell.value for cell in first_row}

            columns_difference = expected_column_names.difference(column_names)
            if columns_difference:
                return [
                    {
                        "row_number": 1,
                        "header": col,
                        "message": "Invalid column name",
                    }
                    for col in columns_difference
                ]
            return []

    @classmethod
    def validate_household_rows(cls, *args, **kwargs):
        if cls.WB is None:
            xlsx_file = kwargs.get("file")
            wb = openpyxl.load_workbook(xlsx_file)
        else:
            wb = cls.WB
        household_sheet = wb["Households"]
        return cls.rows_validator(household_sheet)

    @classmethod
    def validate_individuals_rows(cls, *args, **kwargs):
        if cls.WB is None:
            xlsx_file = kwargs.get("file")
            wb = openpyxl.load_workbook(xlsx_file)
        else:
            wb = cls.WB
        household_sheet = wb["Individuals"]
        return cls.rows_validator(household_sheet)
