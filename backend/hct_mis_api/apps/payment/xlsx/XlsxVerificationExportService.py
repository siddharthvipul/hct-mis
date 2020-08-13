from typing import List, Tuple, Dict

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from payment.models import PaymentVerification


class XlsxVerificationExportService:
    HEADERS = (
        "payment_record_id",
        "verification_status",
        "head_of_household",
        "household_id",
        "delivered_amount",
        "received_amount",
    )
    VERIFICATION_SHEET = "Payment Verifications"
    META_SHEET = "Meta"
    VERSION_CELL_NAME_COORDINATES = "A1"
    VERSION_CELL_COORDINATES = "B1"
    VERSION_CELL_NAME = "FILE_TEMPLATE_VERSION"
    VERSION = "1.0"

    def __init__(self, cashplan_payment_verification):
        self.cashplan_payment_verification = cashplan_payment_verification
        self.payment_record_verifications = (
            cashplan_payment_verification.payment_record_verifications.all()
        )

    def _create_workbook(self) -> openpyxl.Workbook:
        wb = openpyxl.Workbook()
        ws_verifications = wb.active
        ws_verifications.title = XlsxVerificationExportService.VERIFICATION_SHEET
        self.wb = wb
        self.ws_verifications = ws_verifications
        self.ws_meta = wb.create_sheet(XlsxVerificationExportService.META_SHEET)
        return wb

    def _add_version(self):
        self.ws_meta[
            XlsxVerificationExportService.VERSION_CELL_NAME_COORDINATES
        ] = XlsxVerificationExportService.VERSION_CELL_NAME
        self.ws_meta[
            XlsxVerificationExportService.VERSION_CELL_COORDINATES
        ] = XlsxVerificationExportService.VERSION

    def _add_headers(self):
        headers_row = XlsxVerificationExportService.HEADERS
        self.ws_verifications.append(headers_row)

    def _add_payment_record_verification_row(self, payment_record_verification):

        payment_record_verification_row = (
            str(payment_record_verification.payment_record_id),
            payment_record_verification.status,
            str(
                payment_record_verification.payment_record.household.head_of_household.full_name
            ),
            str(payment_record_verification.payment_record.household_id),
            payment_record_verification.payment_record.delivered_quantity,
            payment_record_verification.received_amount,
        )
        self.ws_verifications.append(payment_record_verification_row)

    def _add_payment_record_verifications(self):
        for payment_record_verification in self.payment_record_verifications:
            self._add_payment_record_verification_row(
                payment_record_verification
            )

    def _add_data_validation(self):
        statuses = [x[0] for x in PaymentVerification.STATUS_CHOICES]
        self.dv_verification_status = DataValidation(
            type="list", formula1=f'"{",".join(statuses)}"', allow_blank=False
        )
        self.dv_verification_status.add(
            f"B2:B{len(self.ws_verifications['B'])}"
        )
        self.ws_verifications.add_data_validation(self.dv_verification_status)

    def generate_workbook(self):
        self._create_workbook()
        self._add_version()
        self._add_headers()
        self._add_payment_record_verifications()
        self._add_data_validation()
        self._adjust_column_width_from_col(self.ws_verifications, 0, 1, 5)
        return self.wb

    def generate_file(self, filename):
        self.generate_workbook()
        self.wb.save(filename=filename)

    def _adjust_column_width_from_col(self, ws, min_row, min_col, max_col):

        column_widths = []

        for i, col in enumerate(
            ws.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row)
        ):

            for cell in col:
                value = cell.value
                if value is not None:

                    if isinstance(value, str) is False:
                        value = str(value)

                    try:
                        column_widths[i] = max(column_widths[i], len(value))
                    except IndexError:
                        column_widths.append(len(value))

        for i, width in enumerate(column_widths):
            col_name = get_column_letter(min_col + i)
            value = column_widths[i] + 2
            ws.column_dimensions[col_name].width = value
